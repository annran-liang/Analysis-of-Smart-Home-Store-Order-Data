1.数据预处理
import pandas as pd
# 加载数据
df1 = pd.read_excel("C:\\Users\\华硕\\Desktop\三创赛\\题目数\\202302-202304.xlsx")
df2 = pd.read_excel("C:\\Users\\华硕\\Desktop\三创赛\\题目数据\\202211-202301.xlsx")
df3 = pd.read_excel("C:\\Users\\华硕\\Desktop\三创赛\\题目数据\\202208-202310.xlsx")
df4 = pd.read_excel("C:\\Users\\华硕\\Desktop\三创赛\\题目数据\\202205-202307.xlsx")
# 根据列名合并表格，忽略列的原始排列顺序
merged_df = pd.concat([df1, df2, df3,df4], axis=0, ignore_index=True)
# 合并后的DataFrame保存为新的Excel文件
merged_df.to_excel("C:\\Users\\华硕\\Desktop\\三创赛\\题目数据",index=False)

# 读取Excel文件
file_path = "C:\\Users\\华硕\\Desktop\三创赛\\题目数据\\合并数据.xlsx"
df = pd.read_excel(file_path)

# 定义一个函数来去除"US $"并只保留数字
def remove_us_dollar(value):
    if isinstance(value, str):
        return value.replace('US $', '').strip()
    return value
# 应用这个函数到D列和O列
df['D'] = df['D'].apply(remove_us_dollar)
df['O'] = df['O'].apply(remove_us_dollar)
df['S'] = df['S'].apply(remove_us_dollar)
df['W'] = df['W'].apply(remove_us_dollar)
df['Y'] = df['Y'].apply(remove_us_dollar)

df['V'] = df['V'].replace({“等待买家付款”: 1, “等待买家收货”: 2,”等待您发货”:3, 
“订单关闭”: 4,”冻结中”: 5, “交易完成”: 6,”买家申请取消订单”:7})
df['N'] = df['N'].replace({“平台已税”: 1,”平台未税”: 0,”null”: 2})

data = pd.read_excel(file_path)

# 将 "付款时间 "和 "发货时间 "列转换为日期时间格式
# # 使用 errors='coerce' 来处理转换错误，将其转换为 NaT
data['付款时间'] = pd.to_datetime(data['付款时间'], format='%Y-%m-%d %H:%M', errors='coerce')
data['发货时间'] = pd.to_datetime(data['发货时间'], format='%Y-%m-%d %H:%M', errors='coerce')

#以小时计算差异
data['用时(小时)'] = (data['发货时间'] - data['付款时间']).dt.total_seconds() / 3600

import pandas as pd
# 定义一个函数，以根据指定的条件分配值
def assign_value_corrected(row):
    if pd.isnull(row['发货时间']) or pd.isnull(row['发货期限']):
        return 2  # One of the dates is missing
    elif row['发货时间'] <= row['发货期限']:
        return 0  # Not overdue
    else:
        return 1  # Overdue
# 将该函数应用于每行，以创建新的列
new_data['状态'] = new_data.apply(assign_value_corrected, axis=1)
2.店铺运营情况分析
import pandas as pd
from collections import defaultdict
import matplotlib.pylab as plt
from openpyxl import Workbook

# 分别获取月份和季节的数据
data = pd.read_csv('data.csv', usecols=['下单日期', '订单状态'], parse_dates=True)
month_dict = defaultdict(int)
season_dict = defaultdict(int)
day_dict = defaultdict(int)

for _, row in data.iterrows():
    status = int(row[1])
    # 剔除异常数据
    if status in [4, 5, 7]:
        continue
    date = row[0]
    date_data = date.split('/')
    month = int(date_data[1])
    month_dict[month] += 1
    # 更改数字以具体分析月份
    if month == 4:
        day = int(date_data[2])
        day_dict[day] += 1

for k, v in month_dict.items():
    season_dict[(k-1)//3] += v

# 取消注释代码以分别作图
# 绘制月份销量柱状图
# month_list = month_dict.items()
# month_list = sorted(month_list)
# x, y = zip(*month_list)
# plt.bar(x=x, height=y, width=0.4)
# plt.show()

# 绘制季度销量柱状图
# labels = ['spring', 'summer', 'autumn', 'winter']
# season_list = season_dict.items()
# season_list = sorted(season_list)
# x, y = zip(*season_list)
# plt.xticks(x, labels=labels)
# plt.bar(x, y)
# plt.show()

# 绘制日度销量柱状图
# day_list = day_dict.items()
# day_list = sorted(day_list)
# x, y = zip(*day_list)
# plt.bar(x=x, height=y, width=0.4)
# plt.show()

# wb = Workbook()
# ws = wb.active
# ws.title = "四月日度分析"
# ws.cell(row=1, column=1, value='日期')
# ws.cell(row=1, column=2, value='销量')
# row = 2
# day_list = day_dict.items()
# day_list = sorted(day_list)
# for item in day_list:
#     ws.cell(row=row, column=1, value=item[0])
#     ws.cell(row=row, column=2, value=item[1])
#    row += 1
# wb.save('四月日度分析.xlsx')

import pandas as pd
from collections import defaultdict
import matplotlib.pylab as plt
import matplotlib.pyplot as pyplot
from openpyxl import Workbook

data = pd.read_csv('new_data.csv', usecols=['下单日期' ,'商品信息' ,'数量'], parse_dates=True)
product_dict = defaultdict(int)

for _, row in data.iterrows():
    # 异常数据已剔除
    try:
        split_data = row[1].split('(')
        name = split_data[0].strip()
        num = int(row[2])
    except:
        continue
    product_dict[name] += num

# 得到销量最高的十个产品
product_dict = sorted(product_dict.items(), key = lambda kv:(kv[1], kv[0]), reverse=True)
top_product_list = product_dict[ : 10]

# 给相应产品分配产品编码，便于绘图
code_dict = defaultdict(str)
for i, item in enumerate(top_product_list):
    code = chr(ord('A') + i)
    code_dict[code] = item[0]

# 绘制热销产品柱状图
# x, y = zip(*top_code_list)
# plt.bar(x=x, height=y, width=0.4)
# plt.show()

# 分析热销商品的季节性和属性
# 修改该变量以分析不同产品
selected_product = 3
month_dict = defaultdict(int)
country_dict = defaultdict(int)
color_dict = defaultdict(int)
day_dict = defaultdict(int)
for _, row in data.iterrows():
    try:
        split_data = row[1].split('(')
        name = split_data[0].strip()
    except:
        continue
    if name == top_product_list[selected_product][0]:
        date = row[0].split('/')
        month = int(date[1])
        month_dict[month] += 1
        property = split_data[1][ : -2].split(':')
        country = property[-1]
        country_dict[country] += 1
        if selected_product == 0:
            color = property[2].split()[1].split('、')[0]
        elif selected_product == 1:
            color = property[2].split('、')[0]
        color_type = property[2].split('、')[0]
        color_dict[color_type] += 1
        # 分析日度销量
        if month == 1:
            day = int(date[2])
            day_dict[day] += 1

# wb = Workbook()
# ws = wb.active
# ws.title = "D商品月度分析"
# ws.cell(row=1, column=1, value='月份')
# ws.cell(row=1, column=2, value='销量')
# row = 2
# month_list = month_dict.items()
# month_list = sorted(month_list)
# for item in month_list:
#     ws.cell(row=row, column=1, value=item[0])
#     ws.cell(row=row, column=2, value=item[1])
#     row += 1
# wb.save('D商品月度分析.xlsx')

# 绘制国家销量柱状图
# country_list = country_dict.items()
# x, y = zip(*country_list)
# plt.bar(x=x, height=y, width=0.4)
# plt.title(top_product_list[selected_product][0])
# plt.show()

# 绘制月份销量柱状图
# month_list = month_dict.items()
# month_list = sorted(month_list)
# x, y = zip(*month_list)
# plt.bar(x=x, height=y, width=0.4)
# plt.title(top_product_list[selected_product][0])
# plt.show()

# wb = Workbook()
# ws = wb.active
# ws.title = "D一月日度分析"
# ws.cell(row=1, column=1, value='日期')
# ws.cell(row=1, column=2, value='销量')
# row = 2
# day_list = day_dict.items()
# day_list = sorted(day_list)
# print(day_list)
# for item in day_list:
#     ws.cell(row=row, column=1, value=item[0])
#     ws.cell(row=row, column=2, value=item[1])
#     row += 1
# wb.save('D一月日度分析.xlsx')
3.绘图
import pandas as pd
import matplotlib.pyplot as plt
import re
# 读取Excel文件数据
data = pd.read_excel(r'D:\桌面\数据预处理后.xlsx')
plt.rcParams['font.sans-serif'] = ['SimSun']  # 指定使用宋体字体
plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题

# 创建一个新的列来存储产品名称
data['产品名称'] = ""

# 遍历产品信息，解析产品名称
for index, row in data.iterrows():
    # 根据产品信息的结构提取产品名称
    product_info = row['商品信息']
    product_name = re.search(r'^([^()]*)', product_info).group(1)

    # 将提取到的产品名称存储到新的列中
    data.at[index, '产品名称'] = product_name.strip()

# 将下单日期列转换为日期类型
data['下单日期'] = pd.to_datetime(data['下单日期'], format='%Y/%m/%d')

# 分析下单日期数据，找出电商大促日期
daily_order_count = data['下单日期'].dt.date.value_counts().sort_index()

# 根据实际情况确定大促日期的条件
promotion_threshold = 300  # 假设每日订单数超过300单的日期为大促日期

promotion_dates = daily_order_count[daily_order_count > promotion_threshold].index.tolist()
# 打印电商大促日期
# print("电商大促日期：", promotion_dates)

# 筛选出大促节日的订单
promotion_data = data[data['下单日期'].isin(promotion_dates)]

# 进行促销产品分析，计算促销产品的销售数量
promotion_product_analysis = promotion_data.groupby('产品名称').agg({'产品名称': 'count'}).rename(
    columns={'产品名称': '销售数量'}).reset_index()

# 按销售数量降序排序
promotion_product_analysis = promotion_product_analysis.sort_values('销售数量', ascending=False)
# 仅保留前10个产品
promotion_product_analysis = promotion_product_analysis.head(10)
# 保存为Excel文件
promotion_product_analysis.to_excel(r'C:\Users\87202\Desktop\季节性产品分析\promotion_product_analysis.xlsx', index=False)
# 保存为Excel文件
promotion_dates_count = daily_order_count[daily_order_count.index.isin(promotion_dates)].head(10).reset_index()
promotion_dates_count.columns = ['日期', '销售数量']
promotion_dates_count.to_excel(r'C:\Users\87202\Desktop\季节性产品分析\promotion_dates_count.xlsx', index=False)
# 创建包含两个子图的图表
fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))

# 绘制柱状图
ax1.bar(promotion_product_analysis['产品名称'], promotion_product_analysis['销售数量'])
ax1.set_xticks(range(len(promotion_product_analysis['产品名称'])))  # 设置刻度位置
ax1.set_xticklabels(promotion_product_analysis['产品名称'], rotation=90)
ax1.set_xlabel('产品名称')
ax1.set_ylabel('销售数量')
ax1.set_title('前10个大促产品销售数量')

# 绘制折线图
ax2.plot(range(len(promotion_dates_count['日期'])), promotion_dates_count['销售数量'], linestyle='-', marker='o', color='orange')
ax2.set_xticks(range(len(promotion_dates_count['日期'])))  # 设置刻度位置
ax2.set_xticklabels(promotion_dates_count['日期'], rotation=45)
ax2.set_xlabel('日期')
ax2.set_ylabel('销售数量')
ax2.set_title('前10个大促日期销售数量')

# 调整子图之间的间距
plt.subplots_adjust(wspace=0.3)
plt.show()

import pandas as pd
import matplotlib.pyplot as plt
import re
# 读取Excel文件数据
data = pd.read_excel(r'C:\Users\87202\Desktop\季节性产品分析\数据预处理后.xlsx')
plt.rcParams['font.sans-serif'] = ['SimSun']  # 指定使用宋体字体
plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
# 删除不必要的空格
data['州/省'] = data['州/省'].str.strip()
# 将产品总金额列转换为数值类型
data['产品总金额'] = data['产品总金额'].astype(float)
# 按州/省分组并计算销售额之和
state_sales = data.groupby('州/省')['产品总金额'].sum()
# 排序并选择销售额前10的州/省
top_10_states = state_sales.sort_values(ascending=False).head(10)
# 保存为Excel文件
top_10_states.to_excel(r'C:\Users\87202\Desktop\季节性产品分析\top_10_states_analysis.xlsx', index=True)

# 绘制柱状图
plt.figure(figsize=(16, 14))  # 调整图表尺寸
plt.bar(top_10_states.index, top_10_states.values)
plt.xlabel('州/省')
plt.ylabel('销售额')
plt.title('销售额前10的州/省')
plt.xticks(rotation=45)
plt.tight_layout()
plt.show()



import pandas as pd
import matplotlib.pyplot as plt
import re

# 读取Excel文件数据
data = pd.read_excel(r'C:\Users\87202\Desktop\季节性产品分析\数据预处理后.xlsx')
plt.rcParams['font.sans-serif'] = ['SimSun']  # 指定使用宋体字体
plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题

# 最频繁使用的前10个物流公司
# 去除物流公司字段中的空值
data = data.dropna(subset=['买家选择物流'])

# 删除不必要的空格
data['买家选择物流'] = data['买家选择物流'].str.strip()

# 计算每个物流公司的使用频率
logistics_freq = data['买家选择物流'].value_counts().head(10)

# 创建一个DataFrame保存物流公司名称和使用频率
df_logistics_freq = pd.DataFrame({'物流公司名称': logistics_freq.index, '使用频率': logistics_freq.values})

# 保存为Excel文件
df_logistics_freq.to_excel(r'C:\Users\87202\Desktop\季节性产品分析\logistics_freq_analysis.xlsx', index=False)

# 绘制柱状图
plt.figure(figsize=(16, 15))  # 调整图表尺寸
plt.bar(df_logistics_freq['物流公司名称'], df_logistics_freq['使用频率'])
plt.xlabel('物流公司')
plt.ylabel('使用频率')
plt.title('最频繁使用的前10个物流公司')
plt.xticks(rotation=45)
plt.tight_layout()
plt.show()

import pandas as pd
import matplotlib.pyplot as plt
import re

# 读取Excel文件数据
data = pd.read_excel(r'C:\Users\87202\Desktop\季节性产品分析\数据预处理后.xlsx')
plt.rcParams['font.sans-serif'] = ['SimSun']  # 指定使用宋体字体
plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题

# 订单量最高的前10个国家
# 删除不必要的空格
data['收货国家'] = data['收货国家'].str.strip()
# 计算每个国家的订单数量
country_orders = data['收货国家'].value_counts().head(10)
# 创建一个DataFrame保存国家名称和订单数量
df_country_orders = pd.DataFrame({'国家名称': country_orders.index, '订单数量': country_orders.values})
# 保存订单数量到Excel
df_country_orders.to_excel(r'C:\Users\87202\Desktop\季节性产品分析\country_orders_quantity.xlsx', index=False)
# 设置图表大小
plt.figure(figsize=(14, 10))

# 绘制柱状图
plt.bar(df_country_orders['国家名称'], df_country_orders['订单数量'])
plt.xlabel('国家')
plt.ylabel('订单数量')
plt.title('订单量最高的前10个国家')
plt.xticks(rotation=45)
plt.tight_layout()
plt.show()

import pandas as pd
import matplotlib.pyplot as plt

# 读取Excel文件数据
data = pd.read_excel(r'C:\Users\87202\Desktop\季节性产品分析\数据预处理后.xlsx')
plt.rcParams['font.sans-serif'] = ['SimSun']  # 指定使用宋体字体
plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题

# 将下单日期和付款时间转换为DateTime类型
data['下单日期'] = pd.to_datetime(data['下单日期'])
data['付款时间'] = pd.to_datetime(data['付款时间'])

# 提取下单日期中的月份作为季节
data['季节'] = data['下单日期'].dt.month.map({
    1: '冬季', 2: '冬季', 3: '春季',
    4: '春季', 5: '春季', 6: '夏季',
    7: '夏季', 8: '夏季', 9: '秋季',
    10: '秋季', 11: '秋季', 12: '冬季'
})

# 计算每个季节的销量和销售总额
season_sales = data.groupby('季节').agg({'订单金额': sum})

# 保存季节销售总额到Excel
season_sales.to_excel(r'C:\Users\87202\Desktop\季节性产品分析\season_sales_analysis.xlsx')

# 绘制销售总额柱状图
fig, ax = plt.subplots()
season_sales['订单金额'].plot(kind='bar', ax=ax, color='r', width=0.4, label='销售总额')

# 设置图表标签
plt.xlabel('季节')
plt.ylabel('销售总额')
plt.title('季节销售总额')
plt.legend()

plt.tight_layout()
plt.show()

import pandas as pd
import matplotlib.pyplot as plt

# 读取Excel文件数据
data = pd.read_excel(r'C:\Users\87202\Desktop\季节性产品分析\数据预处理后.xlsx')
plt.rcParams['font.sans-serif'] = ['SimSun']  # 指定使用宋体字体
plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题

# 将下单日期和付款时间转换为DateTime类型
data['下单日期'] = pd.to_datetime(data['下单日期'])
data['付款时间'] = pd.to_datetime(data['付款时间'])

# 提取下单日期中的月份作为季节
data['季节'] = data['下单日期'].dt.month.map({
    1: '冬季', 2: '冬季', 3: '春季',
    4: '春季', 5: '春季', 6: '夏季',
    7: '夏季', 8: '夏季', 9: '秋季',
    10: '秋季', 11: '秋季', 12: '冬季'
})

# 计算每个季节的销售数量
season_sales = data.groupby('季节').size().reset_index(name='销售数量')

# 保存季节销售数量到Excel
season_sales.to_excel(r'C:\Users\87202\Desktop\季节性产品分析\season_sales_quantity.xlsx', index=False)

# 绘制销售数量柱状图
fig, ax = plt.subplots()
season_sales.plot(kind='bar', x='季节', y='销售数量', ax=ax, color='b', width=0.4, label='销售数量')

# 设置图表标签
plt.xlabel('季节')
plt.ylabel('销售数量')
plt.title('季节销售数量')
plt.legend()

plt.tight_layout()
plt.show()
4.销售预测
# 假设sales_data是一个已经加载的DataFrame
sales_data = pd.read_excel(r'C:\Users\20439\Desktop\all.xlsx', sheet_name='Gateway Analyzer')

# 创建一个包含所有日期的日期范围
date_range = pd.date_range(start='2022-04-30', end='2023-04-30')

# 创建一个空的DataFrame来存储填充后的数据
filled_sales_data = pd.DataFrame(index=date_range, columns=['Sales'])

# 初始化所有销量为0
filled_sales_data['Sales'] = 0

# 遍历sales_data的行，更新销量数据
for index, row in sales_data.iterrows():
    date_obj = row['Date']  # 假设'Date'是包含日期的列名
    sales = row['Sales']  # 假设'Sales'是包含销量的列名
    # 更新销量数据
    filled_sales_data.loc[date_obj, 'Sales'] = sales

# 打印结果
print(filled_sales_data)
import pandas as pd

# 假设filled_sales_data是一个已经填充好的DataFrame
# filled_sales_data = ...

# 设置输出文件的路径
output_file_path = r'C:\Users\20439\Desktop\forecast\Gateway Analyzer.csv'

# 将DataFrame保存为CSV文件
filled_sales_data.to_csv(output_file_path)  # 设置index=False以避免将索引也保存到文件中

# 读取时间数据的格式化
def parser(x):
    return datetime.strptime(x, '%Y/%m/%d')

# 转换成有监督数据
def timeseries_to_supervised(data, lag=1):
    df = DataFrame(data)
    columns = [df.shift(i) for i in range(1, lag + 1)]  # 数据滑动一格，作为input，df原数据为output
    columns.append(df)
    df = concat(columns, axis=1)
    df.fillna(0, inplace=True)
    return df

# 转换成差分数据
def difference(dataset, interval=1):
    diff = list()
    for i in range(interval, len(dataset)):
        value = dataset[i] - dataset[i - interval]
        diff.append(value)
    return Series(diff)


# 逆差分
def inverse_difference(history, yhat, interval=1):  # 历史数据，预测数据，差分间隔
    return yhat + history[-interval]

# 缩放
def scale(train, test):
    # 根据训练数据建立缩放器
    scaler = MinMaxScaler(feature_range=(-1, 1))
    scaler = scaler.fit(train)
    # 转换train data
    train = train.reshape(train.shape[0], train.shape[1])
    train_scaled = scaler.transform(train)
    # 转换test data
    test = test.reshape(test.shape[0], test.shape[1])
    test_scaled = scaler.transform(test)
    return scaler, train_scaled, test_scaled

# 逆缩放
def invert_scale(scaler, X, value):
    new_row = [x for x in X] + [value]
    array = numpy.array(new_row)
    array = array.reshape(1, len(array))
    inverted = scaler.inverse_transform(array)
    return inverted[0, -1]
# fit LSTM来训练数据
def fit_lstm(train, batch_size, nb_epoch, neurons):
    X, y = train[:, 0:-1], train[:, -1]
    X = X.reshape(X.shape[0], 1, X.shape[1])
    model = Sequential()
    # 添加LSTM层
    model.add(LSTM(neurons, batch_input_shape=(batch_size, X.shape[1], X.shape[2]), stateful=True))
    model.add(Dense(1))  # 输出层1个node
    # 编译，损失函数mse+优化算法adam
    model.compile(loss='mean_squared_error', optimizer='adam')
    for i in range(nb_epoch):
        # 按照batch_size，一次读取batch_size个数据
        model.fit(X, y, epochs=1, batch_size=batch_size, verbose=0, shuffle=False)
        model.reset_states()
        print("当前计算次数："+str(i))
    return model

# 1步长预测
def forcast_lstm(model, batch_size, X):
    X = X.reshape(1, 1, len(X))
    yhat = model.predict(X, batch_size=batch_size)
    return yhat[0, 0]

# 加载数据
df = pd.read_csv('C:/Users/20439/Desktop/week/Water Leakage Sensor.csv', header=0, parse_dates=[0], index_col=0)

# 如果您想要将DataFrame转换为Series，可以这样做：
series = df.squeeze()

# 让数据变成稳定的
raw_values = series.values
diff_values = difference(raw_values, 1)#转换成差分数据

# 把稳定的数据变成有监督数据
supervised = timeseries_to_supervised(diff_values, 1)
supervised_values = supervised.values

# 数据拆分：训练数据、测试数据，前40行是训练集，后12行是测试集
train, test = supervised_values[0:-12], supervised_values[-12:]

# 数据缩放
scaler, train_scaled, test_scaled = scale(train, test)

# fit 模型
lstm_model = fit_lstm(train_scaled, 1, 900, 11)  # 训练数据，batch_size，epoche次数, 神经元个数
# 预测
train_reshaped = train_scaled[:, 0].reshape(len(train_scaled), 1, 1)#训练数据集转换为可输入的矩阵
lstm_model.predict(train_reshaped, batch_size=1)#用模型对训练数据矩阵进行预测

# 测试数据的前向验证，实验发现，如果训练次数很少的话，模型回简单的把数据后移，以昨天的数据作为今天的预测值，当训练次数足够多的时候
# 才会体现出来训练结果
predictions = list()
for i in range(len(test_scaled)):#根据测试数据进行预测，取测试数据的一个数值作为输入，计算出下一个预测值，以此类推
    # 1步长预测
    X, y = test_scaled[i, 0:-1], test_scaled[i, -1]
    yhat = forcast_lstm(lstm_model, 1, X)
    # 逆缩放
    yhat = invert_scale(scaler, X, yhat)
    # 逆差分
    yhat = inverse_difference(raw_values, yhat, len(test_scaled) + 1 - i)
    predictions.append(yhat)
    expected = raw_values[len(train) + i + 1]
    print('Day=%d, Predicted=%f, Expected=%f' % (i + 1, yhat, expected))

# 性能报告
rmse = sqrt(mean_squared_error(raw_values[-12:], predictions))
print('Test RMSE:%.3f' % rmse)
# 绘图
import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif'] = ['SimHei']  # 或者 'Microsoft YaHei'
plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题
plt.xlabel('周期数')
plt.ylabel('销量')
plt.title('门窗传感器')
pyplot.plot(predictions)
pyplot.show()

####尝试预测20周期
for i in range(len(test_scaled), len(test_scaled) + 20):
    # 使用上一次的预测值作为新的输入
    X = predictions[-1].reshape(1, 1, -1)
    yhat = forcast_lstm(lstm_model, 1, X)
    # 逆缩放
    yhat = invert_scale(scaler, X, yhat)
    # 逆差分
    yhat = inverse_difference(raw_values, yhat, len(test_scaled) + i)
    predictions.append(yhat)
    print('Day=%d, Predicted=%f' % (i + 1, yhat))
# 输出所有预测结果
print(predictions)
# 绘图
import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif'] = ['SimHei']  # 或者 'Microsoft YaHei'
plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题
plt.xlabel('周期数')
plt.ylabel('销量')
plt.title('门窗传感器')
pyplot.plot(predictions)
pyplot.show()

ef gridsearch_prophet(data, params):
    best_params = None
    best_mse = float('inf')
    # Generate all possible combinations of parameter values
    param_grid = ParameterGrid(params)
    for param in param_grid:
        # Initialize Prophet model with current parameter values
        model = Prophet(seasonality_mode=param['seasonality_mode'],
                        seasonality_prior_scale=param['seasonality_prior_scale'],
                        changepoint_prior_scale=param['changepoint_prior_scale'])
        # Fit the model on the data
        model.fit(data)
        # Make predictions for the next 5 steps
        future = model.make_future_dataframe(periods=5)
        forecast = model.predict(future)
        # Calculate mean squared error
        mse =((data['y']-forecast['yhat'])**2).mean()
        # Update best parameters if current MsE is lower
        if mse < best_mse:
            best_params = param
            best_mse = mse
    return best_params

# Example usage
# data: Pandas DataFrame with 'ds'(date)and 'y'(target)columns
data = pd.read_csv('C:/Users/20439/Desktop/week/Water Leakage Sensor.csv',names=['Date', 'Sales'],skiprows=1, header=None, parse_dates=['Date'], date_parser=lambda x: pd.to_datetime(x, format='%Y-%m-%d'))
data.rename(columns={'Date':'ds','Sales':'y'}, inplace=True)
# 定义参数网格

params={'seasonality_mode':['additive','multiplicative'],
                            'seasonality_prior_scale':[0.01,0.1,1.0],
                            'changepoint_prior_scale':[0.001,8.01,0.1]}
# 执行网格搜索并获取最佳参数
best_params = gridsearch_prophet(data, params)
# 使用最佳参数初始化 Prophet 模型
best_model = Prophet(seasonality_mode=best_params['seasonality_mode'],
                                                  seasonality_prior_scale=best_params['seasonality_prior_scale'],
                                                  changepoint_prior_scale=best_params['changepoint_prior_scale'])
# 在数据上拟合模型
best_model.fit(data)
# 对接下来的 10 个步骤进行预测
future = best_model.make_future_dataframe(periods=10)
forecast = best_model.predict(future)
# 打印接下来 10 个步骤的预测值
print(forecast[['ds','yhat']].tail(10))
forecast[['ds', 'yhat', 'yhat_lower', 'yhat_upper']].tail()
# 展示预测结果
best_model.plot(forecast)
# 预测的成分分析绘图，展示预测中的趋势、周效应和年度效应
best_model.plot_components(forecast)

# 设置字体
chinese_font = FontProperties(fname=r"C:\Windows\Fonts\simsun.ttc", size=12, weight='bold')  # 宋体，加粗
english_font = FontProperties(fname=r"C:\Windows\Fonts\times.ttf", size=10, weight='light')  # Times New Roman，加粗

# 读取CSV文件
data = pd.read_csv('C:/Users/20439/Desktop/预测/Smart Switch.csv')

# 设置图像的DPI
plt.figure(dpi=300)

# 绘制所有数据的销售量
plt.plot(data['Index'], data['Sales'], label='All Sales')

# 获取最后20组数据的索引
last_20_indices = data['Index'].tail(20).index

# 绘制最后20组数据的销售量，用橙色线条突出显示
plt.plot(data.loc[last_20_indices, 'Index'], data.loc[last_20_indices, 'Sales'], color='orange', label='Last 20 Sales')

# 设置标题和轴标签
plt.title('智能开关', fontproperties=chinese_font, fontsize=12,fontweight='bold')
plt.xlabel('周期', fontproperties=chinese_font, fontsize=10,fontweight='bold')
plt.ylabel('销量', fontproperties=chinese_font, fontsize=10,fontweight='bold')

# 设置刻度标签的字体为Times New Roman
plt.yticks(fontproperties = 'Times New Roman', size = 10)
plt.xticks(fontproperties = 'Times New Roman', size = 10)
# 显示图表
plt.show()
